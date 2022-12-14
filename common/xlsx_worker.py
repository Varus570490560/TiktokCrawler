import time

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from common.enums import XlsxWorkerStatus, XlsxReadStatus, XlsxWriteStatus, VideoResponseStatus, XlsxCompareResult
from generate.generate_path import default_path
from common.util import merge_array


class ReadResult:
    status: XlsxReadStatus
    url: str

    def __init__(self, status: XlsxReadStatus, url: str):
        self.status = status
        self.url = url


class ReadPlayedDataResult:
    status: XlsxReadStatus
    datas: dict


# a map from line num to letter
column_dict: dict = {
    1: "A",
    2: "B",
    3: "C",
    4: "D",
    5: "E",
    6: "F",
    7: "G",
    8: "H",
    9: "I",
    10: "J",
    11: "K",
    12: "L",
    13: "M",
    14: "N",
    15: "O",
    16: "P",
    17: "Q",
    18: "R",
    19: "S",
    20: "T",
    21: "U",
    22: "V",
    23: "W",
    24: "X",
    25: "Y",
    26: "Z"
}


class XlsxWorker:
    status: XlsxWorkerStatus
    cur_line: int
    wb: Workbook
    ws: Worksheet
    output_path: str
    column_count: int = 0
    total_rows: int = 0

    def __init__(self):
        pass

    def read_url(self) -> ReadResult:
        if self.status != XlsxWorkerStatus.Reader:
            return ReadResult(XlsxReadStatus.PermissionDenied, "")
        value: str = self.ws.cell(self.cur_line, 1).value
        if value is None:
            return ReadResult(XlsxReadStatus.NoMoreData, "")
        else:
            self.cur_line += 1
            return ReadResult(XlsxReadStatus.Success, value)

    def writer_line(self, value: dict) -> XlsxWriteStatus:
        if self.status != XlsxWorkerStatus.Writer:
            return XlsxWriteStatus.PermissionDenied
        for key in value.keys():
            flag: bool = False
            for i in range(1, self.column_count + 1):
                if self.ws.cell(1, i).value == key:
                    if type(value[key]) in [VideoResponseStatus, XlsxCompareResult]:
                        value[key] = str(value[key].name)
                    self.ws.cell(row=self.cur_line, column=i, value=value[key])
                    flag = True
                    break
            if not flag:
                return XlsxWriteStatus.NoSuchField
        self.cur_line += 1
        self.wb.save(self.output_path)
        return XlsxWriteStatus.Success

    def get_total_rows(self):
        row = 2
        while self.ws.cell(row, 1).value is not None:
            row += 1
            self.total_rows += 1

    def read_line(self) -> ReadPlayedDataResult:
        value: str = self.ws.cell(self.cur_line, 1).value
        if value is None:
            res = ReadPlayedDataResult()
            res.status = XlsxReadStatus.NoMoreData
            res.datas = {}
            return res
        datas = {}
        for column in range(1, 8):
            field = self.ws.cell(1, column).value
            value = self.ws.cell(self.cur_line, column).value
            datas[field] = value
        self.cur_line += 1
        res = ReadPlayedDataResult()
        res.status = XlsxReadStatus.Success
        res.datas = datas
        return res

    # deprecated
    @staticmethod
    def __get_link_unique(column: int, sheet: Worksheet) -> []:
        res = []
        rows = 3
        while True:
            if sheet.max_row == rows:
                return res
            elif sheet.cell(rows, column).value is not None:
                print(sheet.cell(rows, column).value)
                if type(sheet.cell(rows, column).value) is not str:
                    continue
                start = sheet.cell(rows, column).value.find("@") + 1
                unique_id = sheet.cell(rows, column).value[start:]
                if unique_id.find("?") != -1:
                    unique_id = unique_id[0:unique_id.find("?")]
                if unique_id not in res:
                    res.append(unique_id)
            rows += 1

    # deprecated
    @staticmethod
    def __get_unique_id(column: int, sheet: Worksheet) -> []:
        res = []
        rows = 3
        while True:
            if sheet.max_row == rows:
                return res
            elif sheet.cell(rows, column).value is not None:
                unique_id = sheet.cell(rows, column).value
                unique_id = str(unique_id)
                if unique_id.startswith("@"):
                    unique_id = unique_id.replace("@", "")
                if unique_id not in res:
                    res.append(unique_id)
            rows += 1

    @staticmethod
    def __get_link_unique_v2(column: int, sheet: Worksheet) -> set:
        res: set = set()
        rows = 3
        max_row = max((s.row for s in sheet[column_dict[column]] if s.value), default=3)
        print(f"sheet max column: {sheet.max_row}")
        print(f"column max column: {max((s.row for s in sheet[column_dict[column]] if s.value), default=3)}")
        print(column_dict[column])
        print()

        while True:
            if max_row == rows:
                return res
            elif sheet.cell(rows, column).value is not None:
                if type(sheet.cell(rows, column).value) is not str:
                    continue
                start = sheet.cell(rows, column).value.find("@") + 1
                unique_id = sheet.cell(rows, column).value[start:]
                if unique_id.find("?") != -1:
                    unique_id = unique_id[0:unique_id.find("?")]
                res.add(unique_id)
            rows += 1

    @staticmethod
    def __get_unique_id_v2(column: int, sheet: Worksheet) -> set:
        res: set = set()
        max_row = max(list(s.row for s in sheet[column_dict[column]] if s.value), default=3)
        print(f"sheet max column: {sheet.max_row}")
        print(f"column max column: {max((s.row for s in sheet[column_dict[column]] if s.value), default=3)}")
        print(column_dict[column])
        print()
        rows = 3
        while True:
            if max_row == rows:
                return res
            elif sheet.cell(rows, column).value is not None:
                unique_id = sheet.cell(rows, column).value
                unique_id = str(unique_id)
                if unique_id.startswith("@"):
                    unique_id = unique_id.replace("@", "")
                res.add(unique_id)
            rows += 1

    # deprecated
    def read_unique_id(self) -> []:
        res = []
        if self.status != XlsxWorkerStatus.RemoveDupReader:
            return None
        ws = self.wb.worksheets[1]
        res = merge_array(res, self.__get_unique_id(2, ws))
        res = merge_array(res, self.__get_unique_id(11, ws))
        res = merge_array(res, self.__get_link_unique(3, ws))
        res = merge_array(res, self.__get_link_unique(12, ws))
        ws = self.wb.worksheets[2]
        res = merge_array(res, self.__get_unique_id(2, ws))
        res = merge_array(res, self.__get_unique_id(12, ws))
        res = merge_array(res, self.__get_unique_id(20, ws))
        res = merge_array(res, self.__get_link_unique(3, ws))
        res = merge_array(res, self.__get_link_unique(13, ws))
        res = merge_array(res, self.__get_link_unique(21, ws))
        ws = self.wb.worksheets[3]
        res = merge_array(res, self.__get_unique_id(2, ws))
        res = merge_array(res, self.__get_link_unique(3, ws))
        ws = self.wb.worksheets[4]
        res = merge_array(res, self.__get_unique_id(2, ws))
        res = merge_array(res, self.__get_link_unique(3, ws))
        res = merge_array(res, self.__get_unique_id(13, ws))
        res = merge_array(res, self.__get_link_unique(14, ws))
        res = merge_array(res, self.__get_unique_id(19, ws))
        res = merge_array(res, self.__get_link_unique(20, ws))

        for i in range(5, 17):
            ws = self.wb.worksheets[i]
            res = merge_array(res, self.__get_unique_id(1, ws))
        return res

    def read_unique_id_v2(self) -> 'set|None':
        res: set = set()
        if self.status != XlsxWorkerStatus.RemoveDupReader:
            return None
        ws = self.wb.worksheets[1]
        res = self.__get_unique_id_v2(2, ws) | res
        res = self.__get_unique_id_v2(11, ws) | res
        res = self.__get_link_unique_v2(3, ws) | res
        res = self.__get_link_unique_v2(12, ws) | res
        ws = self.wb.worksheets[3]
        res = self.__get_unique_id_v2(2, ws) | res
        res = self.__get_unique_id_v2(12, ws) | res
        res = self.__get_unique_id_v2(20, ws) | res
        res = self.__get_link_unique_v2(3, ws) | res
        res = self.__get_link_unique_v2(13, ws) | res
        res = self.__get_link_unique_v2(21, ws) | res
        ws = self.wb.worksheets[2]
        res = self.__get_unique_id_v2(2, ws) | res
        res = self.__get_link_unique_v2(3, ws) | res
        ws = self.wb.worksheets[4]
        res = self.__get_link_unique_v2(3, ws) | res
        res = self.__get_unique_id_v2(2, ws) | res
        ws = self.wb.worksheets[5]

        res = self.__get_unique_id_v2(2, ws) | res
        res = self.__get_unique_id_v2(11, ws) | res
        res = self.__get_unique_id_v2(17, ws) | res
        res = self.__get_link_unique_v2(3, ws) | res
        res = self.__get_link_unique_v2(12, ws) | res
        res = self.__get_link_unique_v2(18, ws) | res

        for i in range(6, 17):
            ws = self.wb.worksheets[i]
            res = self.__get_unique_id_v2(1, ws) | res
        return res


def init_reader(path: str) -> 'XlsxWorker | None':
    instance = XlsxWorker()
    instance.status = XlsxWorkerStatus.Reader
    try:
        instance.wb = openpyxl.open(filename=path)
        instance.ws = instance.wb.worksheets[0]
    except InvalidFileException:
        return None
    except FileNotFoundError:
        return None
    instance.cur_line = 2
    instance.get_total_rows()
    return instance


def init_writer(path: str, fields: tuple) -> XlsxWorker:
    instance = XlsxWorker()
    instance.wb = openpyxl.Workbook()
    instance.output_path = path
    instance.status = XlsxWorkerStatus.Writer
    instance.ws = instance.wb.worksheets[0]
    column = 1
    instance.column_count = len(fields)
    for field in fields:
        instance.ws.cell(row=1, column=column, value=field)
        column += 1
    instance.wb.save(path)
    instance.cur_line = 2
    return instance


def init_remove_dup_reader() -> 'XlsxWorker | None':
    instance = XlsxWorker()
    instance.status = XlsxWorkerStatus.RemoveDupReader
    try:
        instance.wb = openpyxl.open(filename=default_path + "cache.xlsx")
    except InvalidFileException:
        return None
    return instance


def verify_xlsx_format(path: str) -> bool:
    try:
        openpyxl.open(filename=path)
    except InvalidFileException:
        return False
    return True


def compare(yesterday: XlsxWorker, today: XlsxWorker, name: str):
    writer = init_writer(f"/home/ubuntu/TiktokCrawler/server/history/by_increase/{name}", ("Url", "Status", "VideoId", "Comment", "Share", "Played", "Digg",))
    remove_duplication = []
    while True:
        yesterday_data = yesterday.read_line()
        today.cur_line = 2
        if yesterday_data.status == XlsxReadStatus.NoMoreData:
            break
        remove_duplication.append(yesterday_data.datas["Url"])
        while True:
            today_data = today.read_line()
            if today_data.status == XlsxReadStatus.NoMoreData:
                writer.writer_line({
                    "Url": yesterday_data.datas["Url"],
                    "Status": XlsxCompareResult.TodayNotFound,
                    "VideoId": yesterday_data.datas["VideoId"],
                })
                break
            elif today_data.datas["Url"] != yesterday_data.datas["Url"]:
                continue
            elif today_data.datas["Status"] != "Success" and yesterday_data.datas["Status"] != "Success":
                writer.writer_line({
                    "Url": yesterday_data.datas["Url"],
                    "Status": XlsxCompareResult.BothTwoDaysException,
                    "VideoId": yesterday_data.datas["VideoId"],
                })
            elif today_data.datas["Status"] != "Success":
                writer.writer_line({
                    "Url": yesterday_data.datas["Url"],
                    "Status": XlsxCompareResult.TodayStatusException,
                    "VideoId": yesterday_data.datas["VideoId"],
                })
            elif yesterday_data.datas["Status"] != "Success":
                writer.writer_line({
                    "Url": yesterday_data.datas["Url"],
                    "Status": XlsxCompareResult.YesterdayStatusException,
                    "VideoId": yesterday_data.datas["VideoId"],
                })
            else:
                if type(yesterday_data.datas["Share"]) is None:
                    share_compare = 0
                else:
                    share_compare = int(today_data.datas["Share"]) - int(yesterday_data.datas["Share"])
                writer.writer_line({
                    "Url": yesterday_data.datas["Url"],
                    "Status": XlsxCompareResult.Success,
                    "VideoId": yesterday_data.datas["VideoId"],
                    "Share": share_compare,
                    "Played": int(today_data.datas["Played"]) - int(yesterday_data.datas["Played"]),
                    "Digg": int(today_data.datas["Digg"]) - int(yesterday_data.datas["Digg"]),
                    "Comment": int(today_data.datas["Comment"]) - int(yesterday_data.datas["Comment"]),
                })
    today.cur_line = 2
    while True:
        data = today.read_line()
        if data.status == XlsxReadStatus.NoMoreData:
            break
        elif data.datas["Url"] in remove_duplication:
            continue
        elif data.datas["Status"] == "Success":
            writer.writer_line({
                "Url": data.datas["Url"],
                "Status": XlsxCompareResult.YesterdayNotFound,
                "VideoId": data.datas["VideoId"],
            })
